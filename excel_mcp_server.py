import sys
import os
import json
import re

import uvicorn
from fastapi import FastAPI, Request
from openai import AsyncOpenAI
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv

load_dotenv()

api_key = os.getenv("OPENAI_API_KEY")


# --- Core MCP Classes ---
class BaseAgentTool:
    def __init__(self, name, description, parameters, function):
        self.name = name
        self.description = description
        self.parameters = parameters
        self.function = function

    def run(self, **kwargs):
        return self.function(**kwargs)

    @staticmethod
    def from_function(fn):
        import inspect
        sig = inspect.signature(fn)
        params = {}
        for name, param in sig.parameters.items():
            params[name] = {"type": "string", "description": ""}
        return BaseAgentTool(
            name=fn.__name__,
            description=fn.__doc__ or "",
            parameters=params,
            function=fn
        )

    def openai_tool(self):
        return {
            "type": "function",
            "function": {
                "name": self.name,
                "description": self.description,
                "parameters": {
                    "type": "object",
                    "properties": self.parameters,
                    "required": list(self.parameters.keys())
                }
            }
        }


class ToolFunction:
    @staticmethod
    def from_function(fn):
        return BaseAgentTool.from_function(fn)


class MCPHandler:
    def __init__(self, tool_functions):
        self.tool_functions = tool_functions

    async def acall(self, client, mcp_message):
        tool_call = mcp_message['tool_calls'][0]
        tool_name = tool_call['function']['name']
        tool_args_raw = tool_call['function']['arguments']

        if isinstance(tool_args_raw, str):
            try:
                tool_args = json.loads(tool_args_raw)
            except Exception:
                tool_args = tool_args_raw
        else:
            tool_args = tool_args_raw

        if not isinstance(tool_args, dict):
            raise ValueError(f"Tool arguments must be a dict, got {type(tool_args)}: {tool_args}")

        for tool in self.tool_functions:
            if tool.name == tool_name:
                return tool.run(**tool_args)

        raise Exception(f"No tool found for {tool_name}")

    def openai_tools(self):
        return [tool.openai_tool() for tool in self.tool_functions]


# --- Excel Functions ---

def list_sheets(filepath: str):
    """List all sheet names in an Excel file."""
    wb = load_workbook(filepath)
    return {"sheets": wb.sheetnames}

def read_cell(filepath: str, sheet_name: str, cell: str):
    """Read a value from a specific cell."""
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    value = ws[cell].value
    return {"value": value}

def write_cell(filepath: str, sheet_name: str, cell: str, value: str):
    """Write a value to a specific cell. Create sheet if missing. Handle spaces, cases safely."""
    wb = load_workbook(filepath)
    
    # Match ignoring cases and spaces
    clean_sheet_names = {s.strip().lower(): s for s in wb.sheetnames}
    requested_sheet_clean = sheet_name.strip().lower()
    
    if requested_sheet_clean in clean_sheet_names:
        real_sheet_name = clean_sheet_names[requested_sheet_clean]
        ws = wb[real_sheet_name]
    else:
        # Create sheet if not found
        ws = wb.create_sheet(title=sheet_name)
    
    ws[cell] = value
    wb.save(filepath)
    return {"message": f"Value '{value}' written to {sheet_name}:{cell}"}




def insert_range_in_column(filepath: str, sheet_name: str, start_cell: str, start_value: int, end_value: int):
    """Insert values from start_value to end_value in a column, starting at start_cell."""
    wb = load_workbook(filepath)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # Extract column letter and row number
    match = re.match(r"([A-Za-z]+)([0-9]+)", start_cell.strip())
    if not match:
        return {"error": f"Invalid start_cell format: '{start_cell}'. Must be like 'A2' or 'B5'."}

    col_letter, start_row = match.groups()
    start_row = int(start_row)

    # 🛠 Safely convert start_value and end_value to integers
    try:
        start_value = int(start_value)
        end_value = int(end_value)
    except ValueError:
        return {"error": f"Start value and end value must be integers. Got: start_value={start_value}, end_value={end_value}"}

    for i, value in enumerate(range(start_value, end_value + 1)):
        cell = f"{col_letter}{start_row + i}"
        ws[cell] = value

    wb.save(filepath)
    return {"message": f"Inserted values {start_value} to {end_value} in column {col_letter} starting at {start_cell}."}





def create_sheet(filepath: str, sheet_name: str):
    """Create a new sheet, create file if it doesn't exist."""
    if not os.path.exists(filepath):
        wb = Workbook()
        wb.create_sheet(title=sheet_name)
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)
    else:
        wb = load_workbook(filepath)
        wb.create_sheet(title=sheet_name)
    
    wb.save(filepath)
    return {"message": f"Sheet '{sheet_name}' created in '{filepath}'."}


def delete_sheet(filepath: str, sheet_name: str):
    """Delete a sheet."""
    wb = load_workbook(filepath)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(filepath)
        return {"message": f"Sheet '{sheet_name}' deleted."}
    else:
        return {"error": f"Sheet '{sheet_name}' not found."}

def save_file(filepath: str):
    """Save the Excel file."""
    # Nothing to do because openpyxl.save() happens on write_cell etc.
    return {"message": f"File '{filepath}' saved."}


# --- App Setup ---

# Initialize OpenAI client
openai_client = AsyncOpenAI(api_key=api_key)

# Initialize FastAPI app
app = FastAPI()

# Register MCP Tools
mcp_handler = MCPHandler(
    tool_functions=[
        ToolFunction.from_function(list_sheets),
        ToolFunction.from_function(read_cell),
        ToolFunction.from_function(write_cell),
        ToolFunction.from_function(insert_range_in_column),
        ToolFunction.from_function(create_sheet),
        ToolFunction.from_function(delete_sheet),
        ToolFunction.from_function(save_file),
    ]
)

# Route 1: Direct MCP Call
@app.post("/mcp")
async def handle_mcp(request: Request):
    body = await request.json()
    response = await mcp_handler.acall(openai_client, body)
    return response


@app.post("/ask")
async def ask_excel(request: Request):
    data = await request.json()
    prompt = data.get("prompt")

    default_filepath = "uploaded_file.xlsx"

    chat_response = await openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": (
                    "You are an Excel control agent. "
                    "Use MCP tools only. Respond ONLY in structured MCP tool calls. "
                    "You are allowed to plan and generate multiple tool calls if required to fulfill user requests fully."
                    f"If a tool requires a 'filepath' argument, always use '{default_filepath}'."
                )
            },
            {"role": "user", "content": prompt}
        ],
        tools=mcp_handler.openai_tools(),
        tool_choice="auto",
    )

    tool_calls = chat_response.choices[0].message.tool_calls
    if not tool_calls:
        return {"error": "No MCP tool call detected from model."}

    all_responses = []
    
    for tool_call_obj in tool_calls:
        tool_call = tool_call_obj.to_dict()
        arguments = tool_call["function"]["arguments"]
        if isinstance(arguments, str):
            arguments = json.loads(arguments)

        if "filepath" not in arguments:
            arguments["filepath"] = default_filepath

        tool_call["function"]["arguments"] = arguments

        # Execute tool call
        mcp_response = await mcp_handler.acall(openai_client, {"tool_calls": [tool_call]})
        all_responses.append(mcp_response)

    return {"results": all_responses}



# Main server runner
if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
